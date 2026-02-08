"""
Data Exporter Module

Exports detection results to Excel files for further data analysis.
Each application run creates a new timestamped Excel file.
"""

import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Any
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Exports detection results to Excel files.
    Uses openpyxl for Excel file creation.
    """
    
    def __init__(self, output_dir: str, session_name: Optional[str] = None):
        """
        Initialize the Excel exporter.
        
        Args:
            output_dir: Directory to save Excel files
            session_name: Optional custom session name
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Create unique filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if session_name:
            self.filename = f"detection_results_{session_name}_{timestamp}.xlsx"
        else:
            self.filename = f"detection_results_{timestamp}.xlsx"
        
        self.filepath = self.output_dir / self.filename
        
        # Detection records
        self.records: List[Dict[str, Any]] = []
        self.session_start = datetime.now()
        self.session_info: Dict[str, Any] = {}
        
        # Try to import openpyxl, fallback to CSV if not available
        self.use_excel = True
        try:
            import openpyxl
            self.openpyxl = openpyxl
            logger.info(f"Excel export enabled: {self.filepath}")
        except ImportError:
            self.use_excel = False
            self.filepath = self.filepath.with_suffix('.csv')
            logger.warning("openpyxl not installed. Using CSV format instead.")
            logger.warning("Install with: pip install openpyxl")
    
    def set_session_info(
        self,
        video_source: str,
        model_name: str = "HuggingFace MobileNetV2",
        framework: str = "pytorch",
        **kwargs
    ):
        """
        Set session metadata.
        
        Args:
            video_source: Source of video (camera index or file path)
            model_name: Name of the model used
            framework: Deep learning framework
            **kwargs: Additional metadata
        """
        self.session_info = {
            "session_start": self.session_start.isoformat(),
            "video_source": str(video_source),
            "model_name": model_name,
            "framework": framework,
            **kwargs
        }
    
    def add_detection(
        self,
        frame_number: int,
        timestamp: float,
        prediction: str,
        confidence: float,
        top_predictions: Optional[List[tuple]] = None,
        is_healthy: bool = True,
        **extra_data
    ):
        """
        Add a detection record.
        
        Args:
            frame_number: Current frame number
            timestamp: Time in seconds from video start
            prediction: Predicted class label
            confidence: Prediction confidence (0-1)
            top_predictions: List of (label, confidence) tuples
            is_healthy: Whether prediction indicates healthy crop
            **extra_data: Additional data to store
        """
        record = {
            "frame_number": frame_number,
            "timestamp_sec": round(timestamp, 3),
            "timestamp_formatted": self._format_time(timestamp),
            "prediction": prediction,
            "confidence": round(confidence, 4),
            "confidence_pct": f"{confidence:.1%}",
            "status": "Healthy" if is_healthy else "Disease Detected",
            "detection_time": datetime.now().isoformat(),
        }
        
        # Add top predictions
        if top_predictions:
            for i, (label, conf) in enumerate(top_predictions[:5], 1):
                record[f"top{i}_label"] = label
                record[f"top{i}_confidence"] = round(conf, 4)
        
        # Add any extra data
        record.update(extra_data)
        
        self.records.append(record)
    
    def _format_time(self, seconds: float) -> str:
        """Format seconds as MM:SS.ms"""
        mins = int(seconds // 60)
        secs = seconds % 60
        return f"{mins:02d}:{secs:05.2f}"
    
    def get_summary_stats(self) -> Dict[str, Any]:
        """Calculate summary statistics from recorded detections."""
        if not self.records:
            return {}
        
        total = len(self.records)
        healthy_count = sum(1 for r in self.records if r["status"] == "Healthy")
        disease_count = total - healthy_count
        
        confidences = [r["confidence"] for r in self.records]
        avg_confidence = sum(confidences) / len(confidences)
        min_confidence = min(confidences)
        max_confidence = max(confidences)
        
        # Count by prediction
        prediction_counts = {}
        for r in self.records:
            pred = r["prediction"]
            prediction_counts[pred] = prediction_counts.get(pred, 0) + 1
        
        # Most common prediction
        most_common = max(prediction_counts.items(), key=lambda x: x[1])
        
        return {
            "total_detections": total,
            "healthy_count": healthy_count,
            "disease_count": disease_count,
            "healthy_percentage": round(healthy_count / total * 100, 1),
            "disease_percentage": round(disease_count / total * 100, 1),
            "avg_confidence": round(avg_confidence, 4),
            "min_confidence": round(min_confidence, 4),
            "max_confidence": round(max_confidence, 4),
            "most_common_prediction": most_common[0],
            "most_common_count": most_common[1],
            "unique_predictions": len(prediction_counts),
            "prediction_distribution": prediction_counts,
        }
    
    def save(self) -> str:
        """
        Save all records to Excel/CSV file.
        
        Returns:
            Path to the saved file
        """
        if not self.records:
            logger.warning("No records to save")
            return ""
        
        if self.use_excel:
            return self._save_excel()
        else:
            return self._save_csv()
    
    def _save_excel(self) -> str:
        """Save records to Excel file with multiple sheets."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import PieChart, Reference, BarChart
        
        wb = Workbook()
        
        # ===== SUMMARY SHEET =====
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        title_font = Font(bold=True, size=14)
        
        # Title
        ws_summary["A1"] = "Crop Disease Detection - Session Report"
        ws_summary["A1"].font = Font(bold=True, size=16)
        ws_summary.merge_cells("A1:D1")
        
        # Session info
        row = 3
        ws_summary[f"A{row}"] = "Session Information"
        ws_summary[f"A{row}"].font = title_font
        row += 1
        
        for key, value in self.session_info.items():
            ws_summary[f"A{row}"] = key.replace("_", " ").title()
            ws_summary[f"B{row}"] = str(value)
            row += 1
        
        ws_summary[f"A{row}"] = "Session End"
        ws_summary[f"B{row}"] = datetime.now().isoformat()
        row += 2
        
        # Summary statistics
        stats = self.get_summary_stats()
        ws_summary[f"A{row}"] = "Detection Statistics"
        ws_summary[f"A{row}"].font = title_font
        row += 1
        
        stat_items = [
            ("Total Detections", stats.get("total_detections", 0)),
            ("Healthy Count", stats.get("healthy_count", 0)),
            ("Disease Count", stats.get("disease_count", 0)),
            ("Healthy Percentage", f"{stats.get('healthy_percentage', 0)}%"),
            ("Disease Percentage", f"{stats.get('disease_percentage', 0)}%"),
            ("Average Confidence", f"{stats.get('avg_confidence', 0):.1%}"),
            ("Min Confidence", f"{stats.get('min_confidence', 0):.1%}"),
            ("Max Confidence", f"{stats.get('max_confidence', 0):.1%}"),
            ("Most Common Prediction", stats.get("most_common_prediction", "N/A")),
            ("Unique Predictions", stats.get("unique_predictions", 0)),
        ]
        
        for label, value in stat_items:
            ws_summary[f"A{row}"] = label
            ws_summary[f"B{row}"] = value
            row += 1
        
        row += 2
        
        # Prediction distribution
        ws_summary[f"A{row}"] = "Prediction Distribution"
        ws_summary[f"A{row}"].font = title_font
        row += 1
        
        ws_summary[f"A{row}"] = "Prediction"
        ws_summary[f"B{row}"] = "Count"
        ws_summary[f"C{row}"] = "Percentage"
        for col in ["A", "B", "C"]:
            ws_summary[f"{col}{row}"].font = header_font
            ws_summary[f"{col}{row}"].fill = header_fill
        row += 1
        
        dist_start_row = row
        for pred, count in sorted(stats.get("prediction_distribution", {}).items(), 
                                  key=lambda x: x[1], reverse=True):
            ws_summary[f"A{row}"] = pred.replace("_", " ")
            ws_summary[f"B{row}"] = count
            ws_summary[f"C{row}"] = f"{count / stats['total_detections'] * 100:.1f}%"
            row += 1
        
        # Adjust column widths
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 25
        ws_summary.column_dimensions["C"].width = 15
        
        # ===== DETECTIONS SHEET =====
        ws_detections = wb.create_sheet("Detections")
        
        # Headers
        headers = [
            "Frame", "Timestamp", "Time (sec)", "Prediction", 
            "Confidence", "Status", "Top 2", "Top 2 Conf",
            "Top 3", "Top 3 Conf", "Detection Time"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_detections.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        healthy_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        disease_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        
        for row_idx, record in enumerate(self.records, 2):
            ws_detections.cell(row=row_idx, column=1, value=record.get("frame_number", ""))
            ws_detections.cell(row=row_idx, column=2, value=record.get("timestamp_formatted", ""))
            ws_detections.cell(row=row_idx, column=3, value=record.get("timestamp_sec", ""))
            ws_detections.cell(row=row_idx, column=4, value=record.get("prediction", "").replace("_", " "))
            ws_detections.cell(row=row_idx, column=5, value=record.get("confidence_pct", ""))
            ws_detections.cell(row=row_idx, column=6, value=record.get("status", ""))
            ws_detections.cell(row=row_idx, column=7, value=record.get("top2_label", "").replace("_", " ") if record.get("top2_label") else "")
            ws_detections.cell(row=row_idx, column=8, value=f"{record.get('top2_confidence', 0):.1%}" if record.get("top2_confidence") else "")
            ws_detections.cell(row=row_idx, column=9, value=record.get("top3_label", "").replace("_", " ") if record.get("top3_label") else "")
            ws_detections.cell(row=row_idx, column=10, value=f"{record.get('top3_confidence', 0):.1%}" if record.get("top3_confidence") else "")
            ws_detections.cell(row=row_idx, column=11, value=record.get("detection_time", ""))
            
            # Color code rows
            fill = healthy_fill if record.get("status") == "Healthy" else disease_fill
            for col in range(1, 12):
                ws_detections.cell(row=row_idx, column=col).fill = fill
        
        # Adjust column widths
        col_widths = [8, 12, 10, 35, 12, 16, 30, 12, 30, 12, 22]
        for i, width in enumerate(col_widths, 1):
            ws_detections.column_dimensions[chr(64 + i)].width = width
        
        # Freeze header row
        ws_detections.freeze_panes = "A2"
        
        # ===== RAW DATA SHEET =====
        ws_raw = wb.create_sheet("Raw Data")
        
        if self.records:
            # All columns from records
            all_keys = list(self.records[0].keys())
            
            for col, key in enumerate(all_keys, 1):
                cell = ws_raw.cell(row=1, column=col, value=key)
                cell.font = header_font
                cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            
            for row_idx, record in enumerate(self.records, 2):
                for col, key in enumerate(all_keys, 1):
                    ws_raw.cell(row=row_idx, column=col, value=record.get(key, ""))
        
        # Save workbook
        wb.save(self.filepath)
        logger.info(f"Excel file saved: {self.filepath}")
        
        return str(self.filepath)
    
    def _save_csv(self) -> str:
        """Fallback: Save records to CSV file."""
        import csv
        
        if not self.records:
            return ""
        
        # Get all unique keys
        all_keys = set()
        for record in self.records:
            all_keys.update(record.keys())
        all_keys = sorted(all_keys)
        
        with open(self.filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=all_keys)
            writer.writeheader()
            writer.writerows(self.records)
        
        logger.info(f"CSV file saved: {self.filepath}")
        return str(self.filepath)
    
    def __enter__(self):
        """Context manager entry."""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - auto-save on exit."""
        if self.records:
            self.save()


def create_exporter(output_dir: str, session_name: Optional[str] = None) -> ExcelExporter:
    """
    Factory function to create an Excel exporter.
    
    Args:
        output_dir: Directory to save files
        session_name: Optional session identifier
        
    Returns:
        ExcelExporter instance
    """
    return ExcelExporter(output_dir, session_name)
